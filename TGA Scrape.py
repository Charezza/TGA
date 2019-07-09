from bs4 import BeautifulSoup
import requests
import csv
from subprocess import Popen
import webbrowser
import re
from openpyxl import *
import tkinter as tk
import openpyxl
from itertools import zip_longest


#Variables
#Name of course
CourseName = 'AURAFA008'#input("Input Course Code: ")
#Base URL
TGAURL = 'https://training.gov.au/Training/Details/'
#.csv filename
CourseCSV = CourseName + '.csv'
CourseXLSX = CourseName + '.xlsx'
#Total URL of course
CourseURL = TGAURL + CourseName
#URL get
website_url = requests.get(CourseURL).text
#Beautiful soup work
soup = BeautifulSoup(website_url,'lxml')
table = soup.table
#Excel Frameworks
# wb = Workbook()
wb = openpyxl.Workbook()
ws = wb.active
output_row = 1

#Open URL in browser
#webbrowser.open(CourseURL, 2)
# Define the tables I want to grab
Elements = (soup.find("h2", string="Elements and Performance Criteria")).find_next('table')
Foundation = (soup.find("h2", string="Foundation Skills")).find_next('table')

#Extract the tables

for table_row in table.find_all('tr'):
    cells = table_row.find_all('td')
    row = [[row.text for row in cell.find_all('p')] for cell in cells]

    for row_number, cells in enumerate(zip_longest(*row, fillvalue=""), start=output_row):
        for col_number, value in enumerate(cells, start=1):
            ws.cell(column=col_number, row=row_number, value=value)

    output_row += len(cells)

wb.save('output.xlsx')
Foundation_rows = []

for table_row in Foundation.findAll('tr'):
    columns = table_row.findAll('td')
    output_row = []
    for column in columns:
        sub_rows = column.findAll('p')
        for row in sub_rows:
            output_row.append(row.get_text(separator=' '))
    Foundation_rows.append(output_row)


# Write the tables to .xlsx
Tab0 = (CourseName + 'Elements')
Tab1 = (CourseName + 'Foundation')
ws1 = wb.create_sheet(Tab0)
ws2 = wb.create_sheet(Tab1)

# for row in element_rows:
#     ws1.append(row)
# for row in Foundation_rows:
#     ws2.append(row)
# wb.remove(wb['Sheet'])
# wb.save(CourseXLSX)
p = Popen(CourseXLSX, shell=True)
